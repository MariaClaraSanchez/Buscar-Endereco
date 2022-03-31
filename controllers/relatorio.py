from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class Planilha:
    def __init__(self) -> None:
        """Instancia a classe com as configurações iniciais,
        para criação e ativação do arquivo excel.
        """
        self.arquivo_excel = Workbook()
        self.sheetEndereco = self.arquivo_excel.active

    def formatar_planilha(self) -> None:
        """Função responsável pela formatação da planilha,
            ela coloca a guia com a data atual e renomeia as colunas para 
            receber os novos parâmetros, e coloca quebra de linha.
        """

        self.sheetEndereco['A1'] = "Rua"
        self.sheetEndereco['B1'] = "Bairro"
        self.sheetEndereco['C1'] = "Cidade"
        self.sheetEndereco['D1'] = "CEP"

        self.sheetEndereco['A1'].font = Font(
            size=18, bold=True)
        self.sheetEndereco['B1'].font = Font(
            size=18, bold=True)
        self.sheetEndereco['C1'].font = Font(
            size=18, bold=True)
        self.sheetEndereco['D1'].font = Font(
            size=18, bold=True)

        self.sheetEndereco['A1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetEndereco['B1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetEndereco['C1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetEndereco['D1'].alignment = Alignment(
            horizontal='center', vertical='center')

        self.sheetEndereco.column_dimensions['A'].width = 50
        self.sheetEndereco.column_dimensions['B'].width = 50
        self.sheetEndereco.column_dimensions['C'].width = 50
        self.sheetEndereco.column_dimensions['D'].width = 30

        self.arquivo_excel.save('endereco.xlsx')

    def escreve_dados(self, dados: dict) -> None:
        """Função resonsável por inserir os dados na planilha.
        Args:
            dados_vagas (dict): recebe um dicionário contendo os endereco
        """
        i = 0
        for endereco in dados:
            
            self.sheetEndereco.cell(row=i+2, column=1).value = dados[endereco]['rua']
            self.sheetEndereco.cell(row=i+2, column=2).value = dados[endereco]['bairro']
            self.sheetEndereco.cell(row=i+2, column=3).value = dados[endereco]['cidade']
            self.sheetEndereco.cell(row=i+2, column=4).value = endereco

            self.sheetEndereco[f'A{i+2}'].alignment = Alignment(vertical='center')
            self.sheetEndereco[f'B{i+2}'].alignment = Alignment(vertical='center')
            self.sheetEndereco[f'C{i+2}'].alignment = Alignment(vertical='center')
            self.sheetEndereco[f'D{i+2}'].alignment = Alignment(vertical='center')
            i += 1
        self.arquivo_excel.save('endereco.xlsx')
